import os
from datetime import datetime
from .display_time import NEW_YORK, format_new_york_time

import pandas as pd

from stockscanner.analyst_data import analyst_rating_priority
from stockscanner.ranking import setup_priority
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPORT_FOLDER = "reports"
TOP_RESULTS = 20

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="0F243E")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBTITLE_FONT = Font(color="1F1F1F", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
DARK_GREEN_FILL = PatternFill("solid", fgColor="70AD47")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")
RSI_OVERBOUGHT_FILL = PatternFill("solid", fgColor="4FB52A")
RSI_UPPER_FILL = PatternFill("solid", fgColor="D8EDCC")
RSI_NEUTRAL_FILL = PatternFill("solid", fgColor="BFE7F5")
RSI_LOWER_FILL = PatternFill("solid", fgColor="F8DDCC")
RSI_OVERSOLD_FILL = PatternFill("solid", fgColor="F47732")
RS_STRONG_FILL = PatternFill("solid", fgColor="4FB52A")
RS_UPPER_FILL = PatternFill("solid", fgColor="D8EDCC")
RS_NEUTRAL_FILL = PatternFill("solid", fgColor="BFE7F5")
RS_LOWER_FILL = PatternFill("solid", fgColor="F8DDCC")
RS_WEAK_FILL = PatternFill("solid", fgColor="F47732")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def auto_fit_columns(worksheet, maximum_width=35):
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        maximum_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            maximum_length = max(maximum_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(maximum_length + 2, maximum_width)


def format_header_row(worksheet, row_number=1):
    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    worksheet.row_dimensions[row_number].height = 24


def format_data_area(worksheet, start_row=2):
    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def apply_number_formats(worksheet):
    currency_columns = {
        "Current Price",
        "20 MA",
        "50 MA",
        "200 MA",
        "Support Low",
        "Support High",
        "Resistance Low",
        "Resistance High",
        "Zone Tolerance",
        "Entry",
        "Stop Loss",
        "Target 1",
        "Target 2",
        "Target 3",
        "Investment",
    }
    decimal_columns = {
        "RSI",
        "MACD",
        "Relative Strength",
        "Risk/Reward",
    }
    percentage_columns = {
        "Target Upside",
        "Support Distance %",
        "Resistance Distance %",
        "Zone Tolerance %",
    }
    integer_columns = {
        "Rank",
        "Score",
        "Suggested Shares",
        "Support Tests",
        "Resistance Tests",
    }
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    for heading in currency_columns:
        column_number = headers.get(heading)
        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=column_number).number_format = '$#,##0.00'
    for heading in decimal_columns:
        column_number = headers.get(heading)
        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=column_number).number_format = '0.00'
    for heading in percentage_columns:
        column_number = headers.get(heading)
        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=column_number).number_format = '0.00"%"'
    for heading in integer_columns:
        column_number = headers.get(heading)
        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=column_number).number_format = '0'


def apply_recommendation_colors(worksheet):
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    recommendation_column = headers.get("Recommendation")
    if not recommendation_column:
        return
    recommendation_letter = get_column_letter(recommendation_column)
    last_column_letter = get_column_letter(worksheet.max_column)
    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("STRONG BUY",${recommendation_letter}2))'], fill=DARK_GREEN_FILL)
    )
    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(formula=[f'AND(ISNUMBER(SEARCH("BUY",${recommendation_letter}2)),NOT(ISNUMBER(SEARCH("STRONG BUY",${recommendation_letter}2))))'], fill=GREEN_FILL)
    )
    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("ACCUMULATE",${recommendation_letter}2)),ISNUMBER(SEARCH("HOLD",${recommendation_letter}2)))'], fill=YELLOW_FILL)
    )
    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("WATCH",${recommendation_letter}2))'], fill=ORANGE_FILL)
    )
    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("AVOID",${recommendation_letter}2))'], fill=RED_FILL)
    )


def apply_score_colors(worksheet):
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    score_column = headers.get("Score")
    if not score_column:
        return
    score_letter = get_column_letter(score_column)
    score_range = f"{score_letter}2:{score_letter}{worksheet.max_row}"
    worksheet.conditional_formatting.add(score_range, CellIsRule(operator="greaterThanOrEqual", formula=["90"], fill=DARK_GREEN_FILL))
    worksheet.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["70", "89"], fill=GREEN_FILL))
    worksheet.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["50", "69"], fill=YELLOW_FILL))
    worksheet.conditional_formatting.add(score_range, CellIsRule(operator="lessThan", formula=["50"], fill=RED_FILL))


def apply_current_price_rsi_colors(worksheet):
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    current_price_column = headers.get("Current Price")
    rsi_column = headers.get("RSI")
    if not current_price_column or not rsi_column:
        return

    for row_number in range(2, worksheet.max_row + 1):
        rsi = worksheet.cell(row=row_number, column=rsi_column).value
        try:
            rsi = float(rsi)
        except (TypeError, ValueError):
            continue
        if rsi >= 70:
            fill = RSI_OVERBOUGHT_FILL
        elif rsi > 50:
            fill = RSI_UPPER_FILL
        elif rsi == 50:
            fill = RSI_NEUTRAL_FILL
        elif rsi > 30:
            fill = RSI_LOWER_FILL
        else:
            fill = RSI_OVERSOLD_FILL
        worksheet.cell(row=row_number, column=current_price_column).fill = fill


def apply_symbol_relative_strength_colors(worksheet):
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    symbol_column = headers.get("Symbol")
    relative_strength_column = headers.get("Relative Strength")
    if not symbol_column or not relative_strength_column:
        return

    for row_number in range(2, worksheet.max_row + 1):
        relative_strength = worksheet.cell(
            row=row_number, column=relative_strength_column
        ).value
        try:
            relative_strength = float(relative_strength)
        except (TypeError, ValueError):
            continue
        if relative_strength > 5:
            fill = RS_WEAK_FILL
        elif relative_strength > 0:
            fill = RS_LOWER_FILL
        elif relative_strength == 0:
            fill = RS_NEUTRAL_FILL
        elif relative_strength >= -5:
            fill = RS_UPPER_FILL
        else:
            fill = RS_STRONG_FILL
        worksheet.cell(row=row_number, column=symbol_column).fill = fill


def display_current_price_under_symbol(worksheet):
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    symbol_column = headers.get("Symbol")
    current_price_column = headers.get("Current Price")
    if not symbol_column or not current_price_column:
        return

    for row_number in range(2, worksheet.max_row + 1):
        symbol_cell = worksheet.cell(row=row_number, column=symbol_column)
        current_price = worksheet.cell(row=row_number, column=current_price_column).value
        try:
            price_text = f"(${float(current_price):,.2f})"
        except (TypeError, ValueError):
            continue
        symbol_cell.value = CellRichText(
            TextBlock(InlineFont(b=True, sz=11), str(symbol_cell.value)),
            TextBlock(InlineFont(sz=9), f"\n{price_text}"),
        )
        symbol_cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[row_number].height = 30


def prepare_results_dataframe(results):
    df = pd.DataFrame(results)
    if df.empty:
        return df
    numeric_columns = [
        "Current Price",
        "20 MA",
        "50 MA",
        "200 MA",
        "Support Low",
        "Support High",
        "Support Distance %",
        "Resistance Low",
        "Resistance High",
        "Resistance Distance %",
        "Zone Tolerance",
        "Zone Tolerance %",
        "RSI",
        "MACD",
        "Relative Strength",
        "Target Upside",
        "Score",
        "Entry",
        "Stop Loss",
        "Target 1",
        "Target 2",
        "Target 3",
        "Risk/Reward",
        "Suggested Shares",
        "Investment",
    ]
    for column in numeric_columns:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")
    sort_columns = []
    if "Score" in df.columns:
        sort_columns.append("Score")
    if "Signal" in df.columns:
        df["_setup_priority"] = df["Signal"].map(setup_priority)
        sort_columns.append("_setup_priority")
    sort_columns.extend(
        column
        for column in ["Risk/Reward", "Relative Strength"]
        if column in df.columns
    )
    if "Analyst Rating" in df.columns:
        df["_analyst_priority"] = df["Analyst Rating"].map(analyst_rating_priority)
        sort_columns.append("_analyst_priority")
    if "Target Upside" in df.columns:
        df["_target_upside_priority"] = df["Target Upside"].fillna(float("-inf"))
        sort_columns.append("_target_upside_priority")
    if sort_columns:
        df = df.sort_values(by=sort_columns, ascending=[False] * len(sort_columns))
    df = df.drop(
        columns=[
            "Price As Of",
            "_setup_priority",
            "_analyst_priority",
            "_target_upside_priority",
        ],
        errors="ignore",
    )
    df = df.reset_index(drop=True)
    if "Rank" in df.columns:
        df = df.drop(columns=["Rank"])
    df.insert(0, "Rank", range(1, len(df) + 1))
    zone_columns = [
        "Zone Status",
        "Support Low",
        "Support High",
        "Support Distance %",
        "Support Tests",
        "Support Confidence",
        "Support Details",
        "Resistance Low",
        "Resistance High",
        "Resistance Distance %",
        "Resistance Tests",
        "Resistance Confidence",
        "Resistance Details",
        "Zone Tolerance",
        "Zone Tolerance %",
    ]
    present_zone_columns = [column for column in zone_columns if column in df.columns]
    if present_zone_columns:
        remaining = [column for column in df.columns if column not in present_zone_columns]
        insert_at = remaining.index("Current Price") + 1 if "Current Price" in remaining else 1
        ordered = remaining[:insert_at] + present_zone_columns + remaining[insert_at:]
        df = df[ordered]
    return df


def create_dashboard_sheet(workbook, dataframe):
    worksheet = workbook.create_sheet(title="Dashboard", index=0)
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:D2")
    title_cell = worksheet["A1"]
    title_cell.value = "AI Stock Scanner Dashboard"
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    scan_time = format_new_york_time()
    total_stocks = len(dataframe)
    average_score = dataframe["Score"].mean() if "Score" in dataframe.columns else 0
    highest_score = dataframe["Score"].max() if "Score" in dataframe.columns else 0
    recommendation_counts = dataframe["Recommendation"].value_counts() if "Recommendation" in dataframe.columns else pd.Series(dtype=int)
    summary_rows = [
        ("Stocks scanned", total_stocks),
        ("Strong Buy", int(recommendation_counts.get("🟢 STRONG BUY", 0))),
        ("Buy", int(recommendation_counts.get("🟢 BUY", 0))),
        ("Accumulate", int(recommendation_counts.get("🟡 ACCUMULATE", 0))),
        ("Hold", int(recommendation_counts.get("🟡 HOLD", 0))),
        ("Watch", int(recommendation_counts.get("🟠 WATCH", 0))),
        ("Avoid", int(recommendation_counts.get("🔴 AVOID", 0))),
        ("Average score", round(average_score, 2)),
        ("Highest score", round(highest_score, 2)),
    ]
    worksheet["A6"] = "Metric"
    worksheet["B6"] = "Value"
    for cell in worksheet[6][:2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for row_number, (metric, value) in enumerate(summary_rows, start=7):
        worksheet.cell(row=row_number, column=1, value=metric)
        worksheet.cell(row=row_number, column=2, value=value)
        worksheet.cell(row=row_number, column=1).border = THIN_BORDER
        worksheet.cell(row=row_number, column=2).border = THIN_BORDER
    chart_start_row = 7
    worksheet["D6"] = "Recommendation"
    worksheet["E6"] = "Count"
    worksheet["D6"].fill = HEADER_FILL
    worksheet["D6"].font = HEADER_FONT
    worksheet["E6"].fill = HEADER_FILL
    worksheet["E6"].font = HEADER_FONT
    worksheet["G3"] = "Page refreshed as of"
    worksheet["H3"] = scan_time
    worksheet["G3"].font = Font(bold=True, color="1F4E78")
    worksheet["H3"].font = Font(italic=True, color="475569")
    chart_categories = ["Strong Buy", "Buy", "Accumulate", "Hold", "Watch", "Avoid"]
    chart_values = [summary_rows[1][1], summary_rows[2][1], summary_rows[3][1], summary_rows[4][1], summary_rows[5][1], summary_rows[6][1]]
    for row_number, (category, count) in enumerate(zip(chart_categories, chart_values), start=chart_start_row):
        worksheet.cell(row=row_number, column=4, value=category)
        worksheet.cell(row=row_number, column=5, value=count)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Recommendation Breakdown"
    chart.y_axis.title = "Number of Stocks"
    chart.x_axis.title = "Recommendation"
    chart.height = 8
    chart.width = 14
    data = Reference(worksheet, min_col=5, min_row=6, max_row=12)
    categories = Reference(worksheet, min_col=4, min_row=7, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, "G5")
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 12


def export_excel_workbook(filename, dataframe):
    if dataframe.empty:
        print(f"No data available for Excel export: {filename}")
        return None

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        top_opportunities = dataframe.head(TOP_RESULTS)
        top_opportunities.to_excel(writer, sheet_name="Top Opportunities", index=False)
        dataframe.to_excel(writer, sheet_name="Complete Scan", index=False)

    workbook = load_workbook(filename)
    create_dashboard_sheet(workbook, dataframe)

    for sheet_name in ["Top Opportunities", "Complete Scan"]:
        worksheet = workbook[sheet_name]
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        format_header_row(worksheet, row_number=1)
        format_data_area(worksheet, start_row=2)
        apply_number_formats(worksheet)
        apply_recommendation_colors(worksheet)
        apply_score_colors(worksheet)
        apply_current_price_rsi_colors(worksheet)
        apply_symbol_relative_strength_colors(worksheet)
        display_current_price_under_symbol(worksheet)
        auto_fit_columns(worksheet)

    workbook.save(filename)
    return filename


HTML_STYLE = """<style>
body {
    font-family: Arial, sans-serif;
    margin: 0;
    padding: 0;
    background: #f5f7fb;
    color: #0f243e;
    overflow-x: clip;
}
.container {
    max-width: 1200px;
    margin: 0 auto;
    padding: 24px;
}
header {
    margin-bottom: 12px;
}
header h1 {
    margin: 0 0 3px;
    font-size: 1.4rem;
    line-height: 1.2;
}
header p {
    margin: 0;
    color: #475569;
}
nav {
    margin: 8px 0 12px;
}
nav a {
    margin-right: 16px;
    color: #1f4e78;
    text-decoration: none;
    font-weight: 600;
}
nav a:hover {
    text-decoration: underline;
}
.summary,
.report-links {
    margin-bottom: 24px;
}
.summary table,
.report-links table {
    width: calc(100vw - 48px);
    margin-left: calc(50% - 50vw + 24px);
    border-collapse: collapse;
}
.summary th,
.summary td,
.report-links th,
.report-links td {
    border: 1px solid #d9e1f2;
    padding: 10px;
    text-align: left;
}
.summary th,
.report-links th {
    background: #1f4e78;
    color: #ffffff;
}
.stockscanner-table {
    width: calc(100vw - 48px);
    margin-left: calc(50% - 50vw + 24px);
    border-collapse: collapse;
    margin-bottom: 32px;
}
.stockscanner-table th,
.stockscanner-table td {
    border: 1px solid #d9e1f2;
    padding: 10px;
}
.stockscanner-table th {
    background: #1f4e78;
    color: #ffffff;
    position: sticky;
    top: 0;
}
.stockscanner-table tr:nth-child(even) {
    background: #f8fbff;
}
.stockscanner-table td {
    white-space: nowrap;
}
.stockscanner-table td.numeric {
    text-align: right;
}
footer {
    margin-top: 32px;
    padding-top: 16px;
    border-top: 1px solid #d9e1f2;
    color: #475569;
}
@media (max-width: 720px) {
    .container { padding: 12px; }
    .summary table,
    .report-links table,
    .stockscanner-table {
        width: calc(100vw - 24px);
        margin-left: calc(50% - 50vw + 12px);
    }
}
</style>"""


def _prepare_html_dataframe(dataframe):
    if dataframe.empty:
        return dataframe
    df = dataframe.copy()
    numeric_cols = df.select_dtypes(include=["number"]).columns
    if len(numeric_cols) > 0:
        df[numeric_cols] = df[numeric_cols].round(2)
    return df


def _build_html_summary(dataframe):
    total_stocks = len(dataframe)
    average_score = dataframe["Score"].mean() if "Score" in dataframe.columns else 0
    highest_score = dataframe["Score"].max() if "Score" in dataframe.columns else 0
    recommendation_counts = dataframe["Recommendation"].value_counts().to_dict() if "Recommendation" in dataframe.columns else {}
    rows = [
        ("Stocks scanned", total_stocks),
        ("Average score", round(average_score, 2)),
        ("Highest score", round(highest_score, 2)),
        ("Strong Buy", int(recommendation_counts.get("🟢 STRONG BUY", 0)) + int(recommendation_counts.get("STRONG BUY", 0))),
        ("Buy", int(recommendation_counts.get("🟢 BUY", 0)) + int(recommendation_counts.get("BUY", 0))),
        ("Accumulate", int(recommendation_counts.get("🟡 ACCUMULATE", 0)) + int(recommendation_counts.get("ACCUMULATE", 0))),
        ("Hold", int(recommendation_counts.get("🟡 HOLD", 0)) + int(recommendation_counts.get("HOLD", 0))),
        ("Watch", int(recommendation_counts.get("🟠 WATCH", 0)) + int(recommendation_counts.get("WATCH", 0))),
        ("Avoid", int(recommendation_counts.get("🔴 AVOID", 0)) + int(recommendation_counts.get("AVOID", 0))),
    ]
    html_rows = ["<tr><th>Metric</th><th>Value</th></tr>"]
    for metric, value in rows:
        html_rows.append(f"<tr><td>{metric}</td><td>{value}</td></tr>")
    return "<div class='summary'><h2>Summary</h2><table>" + "".join(html_rows) + "</table></div>"


def _build_nav_links(links):
    if not links:
        return ""
    nav_items = [f"<a href='{href}'>{label}</a>" for label, href in links]
    return f"<nav>{''.join(nav_items)}</nav>"


def _create_html_document(title, generated_at, summary_html, table_html, nav_html=None, report_links_html=None):
    return f"""<!DOCTYPE html>
<html lang='en'>
<head>
    <meta charset='utf-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1'>
    <title>{title}</title>
    {HTML_STYLE}
</head>
<body>
    <div class='container'>
        <header>
            <h1>{title}</h1>
            <p>Generated on {generated_at}</p>
        </header>
        {nav_html or ''}
        {report_links_html or ''}
        {summary_html}
        {table_html}
        <footer>
            <p>StockScanner report generated by StockScanner Python package.
            &nbsp;|&nbsp; <a href='/StockScanner/help.html' style='color: #1f4e78;'>Help &amp; FAQ</a></p>
        </footer>
    </div>
</body>
</html>"""


def _write_html_report(filename, dataframe, report_label, nav_links=None):
    html_filename = os.path.splitext(filename)[0] + ".html"
    dataframe = _prepare_html_dataframe(dataframe)
    table_html = dataframe.to_html(
        index=False,
        classes="stockscanner-table",
        border=0,
        justify="left",
        escape=False,
    )
    summary_html = _build_html_summary(dataframe)
    nav_html = _build_nav_links(nav_links)
    generated_at = format_new_york_time()
    title = f"StockScanner {report_label} Report"
    html_content = _create_html_document(title, generated_at, summary_html, table_html, nav_html=nav_html)
    with open(html_filename, "w", encoding="utf-8") as html_file:
        html_file.write(html_content)
    return html_filename


def _write_date_index(date_folder, report_entries):
    index_filename = os.path.join(date_folder, "index.html")
    rows = ["<tr><th>Report</th><th>Type</th><th>File</th></tr>"]
    for entry in report_entries:
        rows.append(
            f"<tr><td>{entry['label']}</td><td>{entry['type']}</td><td><a href='{os.path.basename(entry['path'])}'>{os.path.basename(entry['path'])}</a></td></tr>"
        )
    report_links_html = "<div class='report-links'><h2>Available Reports</h2><table>" + "".join(rows) + "</table></div>"
    generated_at = format_new_york_time()
    display_date = datetime.strptime(os.path.basename(date_folder), "%Y-%m-%d").strftime("%d/%b/%Y")
    title = f"StockScanner Reports for {display_date}"
    html_content = _create_html_document(title, generated_at, "", "", nav_html=None, report_links_html=report_links_html)
    with open(index_filename, "w", encoding="utf-8") as index_file:
        index_file.write(html_content)
    return index_filename


def _write_root_index():
    if not os.path.isdir(REPORT_FOLDER):
        return None
    report_files = []
    for root, _, files in os.walk(REPORT_FOLDER):
        for filename in sorted(files):
            if not filename.lower().endswith(".html"):
                continue
            filepath = os.path.join(root, filename)
            if os.path.normpath(filepath) == os.path.normpath(os.path.join(REPORT_FOLDER, "index.html")):
                continue
            rel_path = os.path.relpath(filepath, REPORT_FOLDER).replace(os.path.sep, "/")
            report_date = os.path.basename(os.path.dirname(filepath))
            report_files.append((report_date, filename, rel_path))

    report_files.sort(key=lambda item: (item[0], item[1]), reverse=True)
    rows = ["<tr><th>Date</th><th>Report</th><th>Link</th></tr>"]
    for report_date, filename, rel_path in report_files:
        display_date = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d/%b/%Y")
        rows.append(
            f"<tr><td>{display_date}</td><td>{filename}</td><td><a href='{rel_path}'>Open</a></td></tr>"
        )
    report_links_html = "<div class='report-links'><h2>All Available Reports</h2><table>" + "".join(rows) + "</table></div>"
    generated_at = format_new_york_time()
    title = "StockScanner Reports Index"
    html_content = _create_html_document(title, generated_at, "", "", nav_html=None, report_links_html=report_links_html)
    root_index = os.path.join(REPORT_FOLDER, "index.html")
    with open(root_index, "w", encoding="utf-8") as root_file:
        root_file.write(html_content)
    return root_index


def export_report(results):
    dataframe = prepare_results_dataframe(results)
    if dataframe.empty:
        print("No data available for Excel export.")
        return None
    scan_date = datetime.now(NEW_YORK).strftime("%Y-%m-%d")
    date_folder = os.path.join(REPORT_FOLDER, scan_date)
    os.makedirs(date_folder, exist_ok=True)
    timestamp = datetime.now(NEW_YORK).strftime("%Y-%m-%d_%H-%M-%S")
    filename = os.path.join(date_folder, f"StockScanner_Combined_{timestamp}.xlsx")
    excel_filename = export_excel_workbook(filename, dataframe)
    html_filename = _write_html_report(excel_filename, dataframe, "Combined")
    report_entries = [
        {"path": excel_filename, "label": "Combined Excel", "type": "Excel"},
        {"path": html_filename, "label": "Combined HTML", "type": "HTML"},
    ]
    _write_date_index(date_folder, report_entries)
    _write_root_index()
    print()
    print("=" * 80)
    print("Professional Excel dashboard created successfully")
    print(excel_filename)
    print(f"HTML report created successfully: {html_filename}")
    print("=" * 80)
    return excel_filename


def export_batch_reports(results, top_count=10, batch_size=50):
    dataframe = prepare_results_dataframe(results)
    if dataframe.empty:
        print("No data available for batch Excel export.")
        return []
    scan_date = datetime.now(NEW_YORK).strftime("%Y-%m-%d")
    date_folder = os.path.join(REPORT_FOLDER, scan_date)
    os.makedirs(date_folder, exist_ok=True)
    timestamp = datetime.now(NEW_YORK).strftime("%Y-%m-%d_%H-%M-%S")
    exported_files = []

    top_df = dataframe.head(top_count)
    report_entries = []
    top_html = None
    if not top_df.empty:
        top_filename = os.path.join(date_folder, f"StockScanner_Top{top_count}_{timestamp}.xlsx")
        export_excel_workbook(top_filename, top_df)
        exported_files.append(top_filename)
        top_html = _write_html_report(top_filename, top_df, f"Top{top_count}")
        report_entries.append({"path": top_filename, "label": f"Top {top_count} Excel", "type": "Excel"})
        report_entries.append({"path": top_html, "label": f"Top {top_count} HTML", "type": "HTML"})

    remainder = dataframe.iloc[top_count:]
    for batch_index, start in enumerate(range(0, len(remainder), batch_size), start=1):
        batch_df = remainder.iloc[start : start + batch_size]
        if batch_df.empty:
            continue
        batch_start = top_count + start + 1
        batch_end = top_count + start + len(batch_df)
        batch_filename = os.path.join(date_folder, f"StockScanner_Batch_{batch_start}-{batch_end}_{timestamp}.xlsx")
        export_excel_workbook(batch_filename, batch_df)
        exported_files.append(batch_filename)
        batch_html = _write_html_report(batch_filename, batch_df, f"Batch {batch_start}-{batch_end}")
        report_entries.append({"path": batch_filename, "label": f"Batch {batch_start}-{batch_end} Excel", "type": "Excel"})
        report_entries.append({"path": batch_html, "label": f"Batch {batch_start}-{batch_end} HTML", "type": "HTML"})

    combined_filename = os.path.join(date_folder, f"StockScanner_Combined_{timestamp}.xlsx")
    export_excel_workbook(combined_filename, dataframe)
    exported_files.append(combined_filename)
    combined_html = _write_html_report(combined_filename, dataframe, "Combined")
    report_entries.append({"path": combined_filename, "label": "Combined Excel", "type": "Excel"})
    report_entries.append({"path": combined_html, "label": "Combined HTML", "type": "HTML"})

    _write_date_index(date_folder, report_entries)
    _write_root_index()

    print()
    print("=" * 80)
    print("Batch Excel and HTML reports created successfully")
    for filename in exported_files:
        print(filename)
    print("=" * 80)

    return exported_files
