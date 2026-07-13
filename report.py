import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =====================================================
# Configuration
# =====================================================

REPORT_FOLDER = "reports"
TOP_RESULTS = 20

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

TITLE_FILL = PatternFill("solid", fgColor="0F243E")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)

SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBTITLE_FONT = Font(color="1F1F1F", bold=True)

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
DARK_GREEN_FILL = PatternFill("solid", fgColor="70AD47")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


# =====================================================
# Utility Functions
# =====================================================

def auto_fit_columns(worksheet, maximum_width=35):
    """Automatically resize worksheet columns."""

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)

        maximum_length = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            maximum_length = max(maximum_length, len(value))

        adjusted_width = min(maximum_length + 2, maximum_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def format_header_row(worksheet, row_number=1):
    """Apply consistent formatting to a header row."""

    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = THIN_BORDER

    worksheet.row_dimensions[row_number].height = 24


def format_data_area(worksheet, start_row=2):
    """Apply borders and alignment to the data section."""

    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column
    ):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center"
            )


def apply_number_formats(worksheet):
    """Apply appropriate formats based on column headings."""

    currency_columns = {
        "Current Price",
        "20 MA",
        "50 MA",
        "200 MA",
        "Entry",
        "Stop Loss",
        "Target 1",
        "Target 2",
        "Target 3",
        "Investment",
    }

    decimal_columns = {
        "RSI",
        "MACD",
        "Relative Strength",
        "Risk/Reward",
    }

    integer_columns = {
        "Rank",
        "Score",
        "Suggested Shares",
    }

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    for heading in currency_columns:
        column_number = headers.get(heading)

        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=row_number,
                    column=column_number
                ).number_format = '$#,##0.00'

    for heading in decimal_columns:
        column_number = headers.get(heading)

        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=row_number,
                    column=column_number
                ).number_format = '0.00'

    for heading in integer_columns:
        column_number = headers.get(heading)

        if column_number:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=row_number,
                    column=column_number
                ).number_format = '0'


def apply_recommendation_colors(worksheet):
    """Color rows based on the recommendation."""

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    recommendation_column = headers.get("Recommendation")

    if not recommendation_column:
        return

    recommendation_letter = get_column_letter(recommendation_column)

    last_column_letter = get_column_letter(worksheet.max_column)

    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[
                f'ISNUMBER(SEARCH("STRONG BUY",${recommendation_letter}2))'
            ],
            fill=DARK_GREEN_FILL
        )
    )

    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[
                f'AND('
                f'ISNUMBER(SEARCH("BUY",${recommendation_letter}2)),'
                f'NOT(ISNUMBER(SEARCH("STRONG BUY",'
                f'${recommendation_letter}2)))'
                f')'
            ],
            fill=GREEN_FILL
        )
    )

    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[
                f'OR('
                f'ISNUMBER(SEARCH("ACCUMULATE",'
                f'${recommendation_letter}2)),'
                f'ISNUMBER(SEARCH("HOLD",'
                f'${recommendation_letter}2))'
                f')'
            ],
            fill=YELLOW_FILL
        )
    )

    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[
                f'ISNUMBER(SEARCH("WATCH",'
                f'${recommendation_letter}2))'
            ],
            fill=ORANGE_FILL
        )
    )

    worksheet.conditional_formatting.add(
        f"A2:{last_column_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[
                f'ISNUMBER(SEARCH("AVOID",'
                f'${recommendation_letter}2))'
            ],
            fill=RED_FILL
        )
    )


def apply_score_colors(worksheet):
    """Apply traffic-light score formatting."""

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    score_column = headers.get("Score")

    if not score_column:
        return

    score_letter = get_column_letter(score_column)
    score_range = f"{score_letter}2:{score_letter}{worksheet.max_row}"

    worksheet.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["90"],
            fill=DARK_GREEN_FILL
        )
    )

    worksheet.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="between",
            formula=["70", "89"],
            fill=GREEN_FILL
        )
    )

    worksheet.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="between",
            formula=["50", "69"],
            fill=YELLOW_FILL
        )
    )

    worksheet.conditional_formatting.add(
        score_range,
        CellIsRule(
            operator="lessThan",
            formula=["50"],
            fill=RED_FILL
        )
    )


def prepare_results_dataframe(results):
    """Convert results into a clean, ranked DataFrame."""

    df = pd.DataFrame(results)

    if df.empty:
        return df

    numeric_columns = [
        "Current Price",
        "20 MA",
        "50 MA",
        "200 MA",
        "RSI",
        "MACD",
        "Relative Strength",
        "Score",
        "Entry",
        "Stop Loss",
        "Target 1",
        "Target 2",
        "Target 3",
        "Risk/Reward",
        "Suggested Shares",
        "Investment",
    ]

    for column in numeric_columns:
        if column in df.columns:
            df[column] = pd.to_numeric(
                df[column],
                errors="coerce"
            )

    sort_columns = [
        column
        for column in [
            "Score",
            "Risk/Reward",
            "Relative Strength"
        ]
        if column in df.columns
    ]

    if sort_columns:
        df = df.sort_values(
            by=sort_columns,
            ascending=[False] * len(sort_columns)
        )

    df = df.reset_index(drop=True)

    if "Rank" in df.columns:
        df = df.drop(columns=["Rank"])

    df.insert(0, "Rank", range(1, len(df) + 1))

    return df


# =====================================================
# Dashboard Creation
# =====================================================

def create_dashboard_sheet(workbook, dataframe):
    """Create the main summary dashboard."""

    worksheet = workbook.create_sheet(
        title="Dashboard",
        index=0
    )

    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells("A1:D2")
    title_cell = worksheet["A1"]

    title_cell.value = "AI Stock Scanner Dashboard"
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    scan_time = datetime.now().strftime(
        "%d %B %Y, %I:%M %p"
    )

    worksheet["A4"] = "Scan timestamp"
    worksheet["B4"] = scan_time

    total_stocks = len(dataframe)

    average_score = (
        dataframe["Score"].mean()
        if "Score" in dataframe.columns
        else 0
    )

    highest_score = (
        dataframe["Score"].max()
        if "Score" in dataframe.columns
        else 0
    )

    recommendation_counts = (
        dataframe["Recommendation"].value_counts()
        if "Recommendation" in dataframe.columns
        else pd.Series(dtype=int)
    )

    summary_rows = [
        ("Stocks scanned", total_stocks),
        (
            "Strong Buy",
            int(
                recommendation_counts.get(
                    "🟢 STRONG BUY",
                    0
                )
            )
        ),
        (
            "Buy",
            int(
                recommendation_counts.get(
                    "🟢 BUY",
                    0
                )
            )
        ),
        (
            "Accumulate",
            int(
                recommendation_counts.get(
                    "🟡 ACCUMULATE",
                    0
                )
            )
        ),
        (
            "Hold",
            int(
                recommendation_counts.get(
                    "🟡 HOLD",
                    0
                )
            )
        ),
        (
            "Watch",
            int(
                recommendation_counts.get(
                    "🟠 WATCH",
                    0
                )
            )
        ),
        (
            "Avoid",
            int(
                recommendation_counts.get(
                    "🔴 AVOID",
                    0
                )
            )
        ),
        ("Average score", round(average_score, 2)),
        ("Highest score", round(highest_score, 2)),
    ]

    worksheet["A6"] = "Metric"
    worksheet["B6"] = "Value"

    for cell in worksheet[6][:2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_number, (metric, value) in enumerate(
        summary_rows,
        start=7
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=metric
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=value
        )

        worksheet.cell(
            row=row_number,
            column=1
        ).border = THIN_BORDER

        worksheet.cell(
            row=row_number,
            column=2
        ).border = THIN_BORDER

    # Recommendation chart data
    chart_start_row = 7

    worksheet["D6"] = "Recommendation"
    worksheet["E6"] = "Count"

    worksheet["D6"].fill = HEADER_FILL
    worksheet["D6"].font = HEADER_FONT
    worksheet["E6"].fill = HEADER_FILL
    worksheet["E6"].font = HEADER_FONT

    chart_categories = [
        "Strong Buy",
        "Buy",
        "Accumulate",
        "Hold",
        "Watch",
        "Avoid",
    ]

    chart_values = [
        summary_rows[1][1],
        summary_rows[2][1],
        summary_rows[3][1],
        summary_rows[4][1],
        summary_rows[5][1],
        summary_rows[6][1],
    ]

    for row_number, (category, count) in enumerate(
        zip(chart_categories, chart_values),
        start=chart_start_row
    ):
        worksheet.cell(
            row=row_number,
            column=4,
            value=category
        )

        worksheet.cell(
            row=row_number,
            column=5,
            value=count
        )

    chart = BarChart()

    chart.type = "col"
    chart.style = 10
    chart.title = "Recommendation Breakdown"
    chart.y_axis.title = "Number of Stocks"
    chart.x_axis.title = "Recommendation"
    chart.height = 8
    chart.width = 14

    data = Reference(
        worksheet,
        min_col=5,
        min_row=6,
        max_row=12
    )

    categories = Reference(
        worksheet,
        min_col=4,
        min_row=7,
        max_row=12
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(categories)

    worksheet.add_chart(
        chart,
        "G5"
    )

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 12


# =====================================================
# Main Export Function
# =====================================================

def export_report(results):
    """Export the scanner output to a formatted Excel workbook."""

    dataframe = prepare_results_dataframe(results)

    if dataframe.empty:
        print("No data available for Excel export.")
        return None

    os.makedirs(
        REPORT_FOLDER,
        exist_ok=True
    )

    timestamp = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )

    filename = os.path.join(
        REPORT_FOLDER,
        f"StockScanner_V4_3_{timestamp}.xlsx"
    )

    top_columns = [
        column
        for column in [
            "Rank",
            "Symbol",
            "Market",
            "Sector",
            "Priority",
            "Score",
            "Signal",
            "Recommendation",
            "Trend",
            "Current Price",
            "Entry",
            "Stop Loss",
            "Target 1",
            "Target 2",
            "Risk/Reward",
            "Relative Strength",
            "Suggested Shares",
            "Investment",
        ]
        if column in dataframe.columns
    ]

    top_opportunities = dataframe.head(
        TOP_RESULTS
    )[top_columns]

    with pd.ExcelWriter(
        filename,
        engine="openpyxl"
    ) as writer:

        top_opportunities.to_excel(
            writer,
            sheet_name="Top Opportunities",
            index=False
        )

        dataframe.to_excel(
            writer,
            sheet_name="Complete Scan",
            index=False
        )

    workbook = load_workbook(filename)

    create_dashboard_sheet(
        workbook,
        dataframe
    )

    for sheet_name in [
        "Top Opportunities",
        "Complete Scan"
    ]:
        worksheet = workbook[sheet_name]

        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        format_header_row(
            worksheet,
            row_number=1
        )

        format_data_area(
            worksheet,
            start_row=2
        )

        apply_number_formats(worksheet)
        apply_recommendation_colors(worksheet)
        apply_score_colors(worksheet)
        auto_fit_columns(worksheet)

    workbook.save(filename)

    print()
    print("=" * 80)
    print("Professional Excel dashboard created successfully")
    print(filename)
    print("=" * 80)

    return filename