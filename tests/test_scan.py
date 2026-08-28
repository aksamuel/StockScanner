import json
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import pytest
from openpyxl import load_workbook

from stockscanner import html_report, report
from stockscanner.add_exception import add_exceptions
from stockscanner.analyst_data import (
    analyst_rating_priority,
    get_analyst_data,
)
from stockscanner.config import MIN_PRICE
from stockscanner.exceptions_dashboard import export_exceptions_dashboard
from stockscanner.html_report import (
    _build_kpi_chart_data,
    _format_new_york_time,
    _generate_html,
)
from stockscanner.market_data import (
    CACHE_DAYS,
    completed_daily_data,
    download_intraday_snapshot,
)
from stockscanner.ranking import rank_stocks, setup_priority
from stockscanner.remove_exception import remove_exception, remove_exceptions
from stockscanner.scan import process_stock
from stockscanner.scoring import score_stock
from stockscanner.signals import generate_signal


def test_score_stock_basic():
    df = pd.DataFrame([
        {
            'Close': 120.0,
            'MA200': 100.0,
            'MA20': 110.0,
            'MA50': 105.0,
            'RSI': 60.0,
            'MACD': 1.0,
            'MACD_SIGNAL': 0.5,
            'AVG_VOLUME': 1000000.0,
            'Volume': 1200000.0,
            'High': 130.0,
        }
    ])
    score = score_stock(df, relative_strength=25)
    assert score >= 50


def test_minimum_share_price_is_one_dollar():
    assert MIN_PRICE == 1.0


def test_daily_cache_is_limited_to_one_day():
    assert CACHE_DAYS == 1


def test_generate_signal_neutral():
    df = pd.DataFrame([
        {
            'Close': 100.0,
            'High': 100.0,
            'MA20': 95.0,
            'MA50': 96.0,
            'MA200': 97.0,
            'RSI': 50.0,
            'MACD': -1.0,
            'MACD_SIGNAL': -0.5,
        }
    ])
    assert generate_signal(df) == "⚪ Neutral"


def test_report_indexes_link_to_dated_reports(tmp_path, monkeypatch):
    monkeypatch.setattr(report, "REPORT_FOLDER", str(tmp_path / "reports"))
    date_folder = tmp_path / "reports" / "2026-08-03"
    date_folder.mkdir(parents=True)

    html_report = date_folder / "StockScanner_Combined_2026-08-03_16-40-23.html"
    html_report.write_text("<html><body>report</body></html>", encoding="utf-8")

    report._write_date_index(
        str(date_folder),
        [{"path": str(html_report), "label": "Combined HTML", "type": "HTML"}],
    )
    report._write_root_index()

    date_index = (date_folder / "index.html").read_text(encoding="utf-8")
    root_index = (tmp_path / "reports" / "index.html").read_text(encoding="utf-8")

    assert "StockScanner Reports for 2026-08-03" in date_index
    assert "href='StockScanner_Combined_2026-08-03_16-40-23.html'" in date_index
    assert "href='2026-08-03/index.html'" in root_index
    assert "href='2026-08-03/StockScanner_Combined_2026-08-03_16-40-23.html'" in root_index


def test_export_exceptions_dashboard(tmp_path):
    exception_list = tmp_path / "exceptions.csv"
    exception_list.write_text(
        "Symbol,Date From,Date To,Reason\n"
        "ABC,01/Aug/2026,31/Aug/2026,Bought & held\n"
        ",,,\n"
        "XYZ,02/Aug/2026,01/Sep/2026,<review>\n",
        encoding="utf-8",
    )
    output = tmp_path / "exceptions.html"

    result = export_exceptions_dashboard(
        str(exception_list),
        str(output),
        "08 August 2026, 04:00 PM",
        today=date(2026, 8, 15),
    )
    page = output.read_text(encoding="utf-8")

    assert result == str(output)
    assert 'content="0; url=my-exceptions.html"' in page
    assert 'window.location.replace("my-exceptions.html")' in page
    assert "Open My Exceptions" in page
    assert "github.com/aksamuel/StockScanner/issues" not in page
    assert "ABC" not in page


def test_legacy_exception_page_does_not_publish_shared_csv_rows(tmp_path):
    exception_list = tmp_path / "exceptions.csv"
    exception_list.write_text(
        "Symbol,Date From,Date To,Reason\n"
        "OLD,01/Jul/2026,31/Jul/2026,Expired\n"
        "LIVE,01/Aug/2026,31/Aug/2026,Active\n",
        encoding="utf-8",
    )
    output = tmp_path / "exceptions.html"

    export_exceptions_dashboard(
        str(exception_list),
        str(output),
        "15 August 2026, 04:00 PM",
        today=date(2026, 8, 15),
    )
    page = output.read_text(encoding="utf-8")

    assert "OLD" not in page
    assert "LIVE" not in page
    assert "my-exceptions.html" in page


def test_remove_exception_matches_complete_symbol_case_insensitively(tmp_path):
    exception_list = tmp_path / "exceptions.csv"
    exception_list.write_text(
        "Symbol,Reason\nABC,First\nABCD,Second\nabc,Duplicate\n",
        encoding="utf-8",
    )

    removed_count = remove_exception("AbC", str(exception_list))
    remaining = exception_list.read_text(encoding="utf-8")

    assert removed_count == 2
    assert "ABC," not in remaining
    assert "abc," not in remaining
    assert "ABCD,Second" in remaining


def test_remove_exceptions_removes_multiple_tickers_atomically(tmp_path):
    exception_list = tmp_path / "exceptions.csv"
    original = "Symbol,Reason\nABC,First\nDEF,Second\nGHI,Third\n"
    exception_list.write_text(original, encoding="utf-8")

    removed_count = remove_exceptions(["ABC", "ghi"], str(exception_list))
    remaining = exception_list.read_text(encoding="utf-8")

    assert removed_count == 2
    assert "DEF,Second" in remaining
    assert "ABC,First" not in remaining
    assert "GHI,Third" not in remaining

    exception_list.write_text(original, encoding="utf-8")
    with pytest.raises(ValueError, match="MISSING"):
        remove_exceptions(["ABC", "MISSING"], str(exception_list))
    assert exception_list.read_text(encoding="utf-8") == original


def test_add_exceptions_adds_thirty_day_rows_atomically(tmp_path):
    exception_list = tmp_path / "exceptions.csv"
    original = "Symbol,Date From,Date To,Reason\nABC,,,Existing\n"
    exception_list.write_text(original, encoding="utf-8")

    added_count = add_exceptions(
        ["def", "GHI"], str(exception_list), date_from=date(2026, 8, 12)
    )
    updated = exception_list.read_text(encoding="utf-8")

    assert added_count == 2
    assert "DEF,12/Aug/2026,11/Sep/2026,Added from scanner dashboard" in updated
    assert "GHI,12/Aug/2026,11/Sep/2026,Added from scanner dashboard" in updated
    assert updated.index("ABC") < updated.index("DEF") < updated.index("GHI")

    custom_list = tmp_path / "custom-exceptions.csv"
    custom_list.write_text(original, encoding="utf-8")
    add_exceptions(
        ["SOC"],
        str(custom_list),
        date_from=date(2026, 8, 12),
        reason="Bought shares",
    )
    assert "SOC,12/Aug/2026,11/Sep/2026,Bought shares" in custom_list.read_text(
        encoding="utf-8"
    )

    with pytest.raises(ValueError, match="ABC"):
        add_exceptions(
            ["JKL", "ABC"], str(exception_list), date_from=date(2026, 8, 12)
        )
    assert "JKL" not in exception_list.read_text(encoding="utf-8")


def test_exception_updates_sort_symbols_alphabetically(tmp_path):
    exception_list = tmp_path / "exceptions.csv"
    exception_list.write_text(
        "Symbol,Date From,Date To,Reason\n"
        "XYZ,,,Last\n"
        "DEF,,,Remove\n"
        "ABC,,,First\n",
        encoding="utf-8",
    )

    remove_exceptions(["DEF"], str(exception_list))
    add_exceptions(["mno"], str(exception_list), date_from=date(2026, 8, 12))
    symbols = pd.read_csv(exception_list)["Symbol"].tolist()

    assert symbols == ["ABC", "MNO", "XYZ"]


def test_scan_dashboard_supports_selecting_top_and_all_results():
    dataframe = pd.DataFrame(
        [
            {
                "Symbol": "ABC",
                "Score": 90,
                "Recommendation": "BUY",
                "Analyst Rating": "Buy",
                "Target Upside": 25,
            },
            {
                "Symbol": "XYZ",
                "Score": 80,
                "Recommendation": "WATCH",
                "Analyst Rating": "Hold",
                "Target Upside": None,
            },
        ]
    )

    page = _generate_html(dataframe, "08 August 2026, 04:00 PM")

    assert page.count('class="exception-select"') == 4
    assert page.count('class="select-all"') == 2
    assert 'data-symbol="ABC"' in page
    assert 'data-symbol="XYZ"' in page
    assert 'id="addExceptions"' in page
    assert 'id="addBought"' in page
    assert "stockscannerPersonalExceptions.add(symbols)" in page
    assert "stockscannerBoughtSelections.add(selections)" in page
    assert "github.com/aksamuel/StockScanner/issues" not in page
    assert "25.00%" in page


def test_current_price_cells_use_rsi_background_colors(tmp_path):
    dataframe = report.prepare_results_dataframe(
        [
            {"Symbol": "HIGH", "Current Price": 10, "RSI": 70, "Score": 90},
            {"Symbol": "UPPER", "Current Price": 11, "RSI": 60, "Score": 80},
            {"Symbol": "NEUTRAL", "Current Price": 12, "RSI": 50, "Score": 70},
            {"Symbol": "LOWER", "Current Price": 13, "RSI": 40, "Score": 60},
            {"Symbol": "LOW", "Current Price": 14, "RSI": 30, "Score": 50},
        ]
    )

    page = _generate_html(dataframe, "12 August 2026, 10:00 PM")
    for css_class in [
        "price-rsi-overbought",
        "price-rsi-upper",
        "price-rsi-neutral",
        "price-rsi-lower",
        "price-rsi-oversold",
    ]:
        assert page.count(f'class="{css_class}"') == 2

    workbook_path = tmp_path / "rsi-colors.xlsx"
    report.export_excel_workbook(workbook_path, dataframe)
    worksheet = load_workbook(workbook_path)["Complete Scan"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    price_column = headers["Current Price"]
    assert [
        worksheet.cell(row=row_number, column=price_column).fill.fgColor.rgb
        for row_number in range(2, 7)
    ] == [
        "004FB52A",
        "00D8EDCC",
        "00BFE7F5",
        "00F8DDCC",
        "00F47732",
    ]


def test_symbol_cells_use_relative_strength_background_colors(tmp_path):
    dataframe = report.prepare_results_dataframe(
        [
            {"Symbol": "STRONG", "Relative Strength": 6, "Score": 90},
            {"Symbol": "UPPER", "Relative Strength": 5, "Score": 80},
            {"Symbol": "NEUTRAL", "Relative Strength": 0, "Score": 70},
            {"Symbol": "LOWER", "Relative Strength": -5, "Score": 60},
            {"Symbol": "WEAK", "Relative Strength": -6, "Score": 50},
        ]
    )

    page = _generate_html(dataframe, "12 August 2026, 10:00 PM")
    for css_class in [
        "symbol-rs-weak",
        "symbol-rs-lower",
        "symbol-rs-neutral",
        "symbol-rs-upper",
        "symbol-rs-strong",
    ]:
        assert page.count(f'class="{css_class}"') == 2

    workbook_path = tmp_path / "relative-strength-colors.xlsx"
    report.export_excel_workbook(workbook_path, dataframe)
    worksheet = load_workbook(workbook_path)["Complete Scan"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    symbol_column = headers["Symbol"]
    assert [
        worksheet.cell(row=row_number, column=symbol_column).fill.fgColor.rgb
        for row_number in range(2, 7)
    ] == [
        "00F47732",
        "00F8DDCC",
        "00BFE7F5",
        "00D8EDCC",
        "004FB52A",
    ]


def test_reports_replace_price_as_of_column_with_refresh_time_near_graph():
    dataframe = report.prepare_results_dataframe(
        [
            {
                "Symbol": "ABC",
                "Price As Of": "2026-08-12T15:15:00-04:00",
                "Score": 90,
            }
        ]
    )

    assert "Price As Of" not in dataframe.columns
    page = _generate_html(
        dataframe,
        "12 August 2026, 10:00 PM",
        page_key="landing",
        chart_data={
            "labels": ["2026-08-10", "2026-08-11"],
            "sp500": [100, 101],
            "top20": [100, 102],
            "constituents": ["ABC"],
            "selected_count": 1,
        },
    )
    marker = "Scan completed: 12 August 2026, 10:00 PM"
    assert page.count(marker) == 1
    assert page.index(marker) < page.index('<canvas id="performanceChart"')
    assert 'id="yahooPriceTime"' in page
    assert 'id="backendRefreshTime"' in page
    assert "Latest Yahoo price:" in page
    assert "Backend price refresh:" in page
    assert "loadDashboardSnapshotTimes();" in page
    assert (
        "Manual Yahoo refreshes run securely on GitHub Actions and redeploy these pages."
        in page
    )
    assert "stockscannerTemporaryRunNewYork" not in page


def test_current_price_is_displayed_below_symbol_in_smaller_text(tmp_path):
    dataframe = report.prepare_results_dataframe(
        [{"Symbol": "ABC", "Current Price": 123.45, "Score": 90}]
    )

    page = _generate_html(dataframe, "12 August 2026, 10:00 PM")
    assert page.count(
        '<span class="symbol-name">ABC</span>'
        '<span class="symbol-price">($123.45)</span>'
    ) == 2

    workbook_path = tmp_path / "symbol-price.xlsx"
    report.export_excel_workbook(workbook_path, dataframe)
    worksheet = load_workbook(workbook_path, rich_text=True)["Complete Scan"]
    headers = {str(cell.value): cell.column for cell in worksheet[1]}
    symbol_cell = worksheet.cell(row=2, column=headers["Symbol"])
    assert str(symbol_cell.value) == "ABC\n($123.45)"
    assert symbol_cell.value[0].font.sz == 11
    assert symbol_cell.value[1].font.sz == 9


def test_scan_time_is_formatted_in_new_york_time():
    utc_time = datetime(2026, 8, 13, 15, 21, tzinfo=ZoneInfo("UTC"))

    assert _format_new_york_time(utc_time) == "13 August 2026, 11:21 AM EDT"


def test_three_report_pages_have_requested_columns_navigation_and_selection():
    dataframe = report.prepare_results_dataframe(
        [
            {
                "Symbol": "ABC",
                "Sector": "Technology",
                "Market": "NYSE",
                "Current Price": 123.45,
                "RSI": 55,
                "Relative Strength": 70,
                "Score": 90,
                "Recommendation": "BUY",
                "Signal": "Strong Uptrend",
                "Trend": "Bullish",
                "MACD": 1.2,
                "Analyst Rating": "Buy",
                "Target Upside": 20,
                "Zone Status": "Between Zones",
                "Suggested Shares": 10,
                "Risk/Reward": 2,
                "Investment": 1234.5,
            }
        ]
    )

    pages = {
        key: _generate_html(dataframe, "12 August 2026, 10:00 PM", page_key=key)
        for key in ["landing", "technical", "analysts", "bought-selection"]
    }
    navigation = {
        "landing": ("index.html", "KPI Dashboard"),
        "technical": ("technical.html", "Technical Analysis"),
        "analysts": ("analysts.html", "Analysts Rating"),
        "bought-selection": ("bought-selection.html", "Bought Selection"),
    }
    for key, page in pages.items():
        current_href, current_label = navigation[key]
        assert (
            f'<span class="nav-current" aria-current="page">{current_label}</span>'
            in page
        )
        assert f'href="{current_href}">{current_label}</a>' not in page
        assert page.count('class="nav-current" aria-current="page"') == 1
        assert 'href="my-exceptions.html">My Exceptions</a>' in page

    technical_headers = pages["technical"].split("<thead><tr>", 1)[1].split(
        "</tr></thead>", 1
    )[0]
    assert "<th>RSI</th>" in technical_headers
    assert "<th>Recommendation</th>" in technical_headers
    assert "<th>Analyst Rating</th>" not in technical_headers
    assert technical_headers.rfind("<th>RSI</th>") < technical_headers.rfind(
        "<th>MACD</th>"
    )
    assert technical_headers.rstrip().endswith("<th>MACD</th>")
    assert 'id="requestYahooRefresh"' in pages["technical"]
    assert '<button id="requestYahooRefresh" type="button">' in pages["technical"]
    assert 'data-current-price="123.45"' in pages["technical"]
    assert 'data-target-one=""' in pages["technical"]
    assert "fetch(snapshotUrl(), { cache: 'no-store' })" in pages["technical"]
    assert "loadRefreshButtonSnapshotTime();" in pages["technical"]
    assert "Refresh Latest Prices · ${snapshot.price_timestamp_new_york}" in pages[
        "technical"
    ]
    assert "row.dataset.currentPrice = price.toString()" in pages["technical"]
    assert "left.dataset.scannerRank" in pages["technical"]
    assert "renumberVisibleRanks" in pages["technical"]
    assert "Could not load the latest price snapshot" in pages["technical"]

    analysts_headers = pages["analysts"].split("<thead><tr>", 1)[1].split(
        "</tr></thead>", 1
    )[0]
    assert "<th>Analyst Rating</th>" in analysts_headers
    assert "<th>Zone Status</th>" in analysts_headers
    assert "<th>MACD</th>" not in analysts_headers

    bought_headers = pages["bought-selection"].split("<thead><tr>", 1)[1].split(
        "</tr></thead>", 1
    )[0]
    assert "<th>Suggested Shares</th>" in bought_headers
    assert "<th>Investment</th>" in bought_headers
    assert pages["technical"].count('class="exception-select"') == 0
    assert pages["analysts"].count('class="exception-select"') == 0
    assert pages["bought-selection"].count('class="exception-select"') == 2
    assert 'id="requestYahooRefresh"' not in pages["analysts"]
    assert 'id="requestYahooRefresh"' not in pages["bought-selection"]
    assert "One-Year Performance: S&amp;P 500 vs Equal-Weight Top 20" in pages["landing"]
    assert "Performance chart unavailable" in pages["landing"]
    assert '<canvas id="performanceChart"' not in pages["landing"]
    assert "width: 100%;\n    max-width: none;" in pages["landing"]
    assert "@media (max-width: 720px)" in pages["landing"]
    assert ".filter-bar input { width: 100%; }" in pages["technical"]
    assert '<table id="topTable">' not in pages["landing"]
    for page_key in ["technical", "analysts", "bought-selection"]:
        assert "One-Year Performance" not in pages[page_key]
        assert '<canvas id="performanceChart"' not in pages[page_key]
        assert "performanceChart" not in pages[page_key]
        assert "Equal-weight Top 20" not in pages[page_key]
        assert "Scan completed:" not in pages[page_key]


def test_analyst_page_symbols_use_current_price_to_support_low_gap_colors():
    dataframe = report.prepare_results_dataframe(
        [
            {"Symbol": "ABOVE5", "Current Price": 100, "Support Low": 94, "Score": 90},
            {"Symbol": "ABOVE0", "Current Price": 100, "Support Low": 95, "Score": 80},
            {"Symbol": "ZERO", "Current Price": 100, "Support Low": 100, "Score": 70},
            {"Symbol": "BELOW0", "Current Price": 100, "Support Low": 105, "Score": 60},
            {"Symbol": "BELOW5", "Current Price": 100, "Support Low": 106, "Score": 50},
        ]
    )

    page = _generate_html(
        dataframe,
        "12 August 2026, 10:00 PM",
        page_key="analysts",
    )

    for css_class in [
        "symbol-support-above-five",
        "symbol-support-above-zero",
        "symbol-support-zero",
        "symbol-support-below-zero",
        "symbol-support-below-five",
    ]:
        assert page.count(f'class="{css_class}"') == 2
    assert 'class="symbol-rs-weak"' not in page


def test_analyst_page_sorts_nearest_support_first_and_missing_support_last():
    dataframe = report.prepare_results_dataframe(
        [
            {"Symbol": "FAR", "Current Price": 100, "Support Low": 80, "Score": 90},
            {"Symbol": "MISSING", "Current Price": 100, "Support Low": None, "Score": 80},
            {"Symbol": "NEAREST", "Current Price": 100, "Support Low": 99, "Score": 70},
            {"Symbol": "NEAR", "Current Price": 100, "Support Low": 95, "Score": 60},
        ]
    )

    page = _generate_html(
        dataframe,
        "12 August 2026, 10:00 PM",
        page_key="analysts",
    )
    top_table = page.split('<table id="topTable">', 1)[1].split("</table>", 1)[0]

    assert top_table.index(">NEAREST</span>") < top_table.index(">NEAR</span>")
    assert top_table.index(">NEAR</span>") < top_table.index(">FAR</span>")
    assert top_table.index(">FAR</span>") < top_table.index(">MISSING</span>")


def test_technical_page_uses_requested_multi_factor_sort_hierarchy():
    dataframe = report.prepare_results_dataframe(
        [
            {
                "Symbol": "LARGEST_GAP",
                "Current Price": 100,
                "Target 1": 130,
                "Relative Strength": -5,
                "Recommendation": "AVOID",
                "Signal": "Neutral",
                "Trend": "Breakout",
                "Score": 90,
            },
            {
                "Symbol": "HIGHER_RS",
                "Current Price": 100,
                "Target 1": 120,
                "Relative Strength": 10,
                "Recommendation": "BUY",
                "Signal": "Pullback to 20 MA",
                "Trend": "Healthy Pullback",
                "Score": 80,
            },
            {
                "Symbol": "STRONGER_REC",
                "Current Price": 100,
                "Target 1": 120,
                "Relative Strength": 5,
                "Recommendation": "STRONG BUY",
                "Signal": "Pullback to 20 MA",
                "Trend": "Healthy Pullback",
                "Score": 70,
            },
            {
                "Symbol": "STRONGER_SIGNAL",
                "Current Price": 100,
                "Target 1": 120,
                "Relative Strength": 5,
                "Recommendation": "BUY",
                "Signal": "Strong Uptrend",
                "Trend": "Healthy Pullback",
                "Score": 60,
            },
            {
                "Symbol": "STRONGER_TREND",
                "Current Price": 100,
                "Target 1": 120,
                "Relative Strength": 5,
                "Recommendation": "BUY",
                "Signal": "Pullback to 20 MA",
                "Trend": "Strong Uptrend",
                "Score": 50,
            },
            {
                "Symbol": "BASE",
                "Current Price": 100,
                "Target 1": 120,
                "Relative Strength": 5,
                "Recommendation": "BUY",
                "Signal": "Pullback to 20 MA",
                "Trend": "Healthy Pullback",
                "Score": 40,
            },
            {
                "Symbol": "LARGER_TIE_GAP",
                "Current Price": 100,
                "Target 1": 125,
                "Relative Strength": 5,
                "Recommendation": "BUY",
                "Signal": "Pullback to 20 MA",
                "Trend": "Healthy Pullback",
                "Score": 35,
            },
            {
                "Symbol": "MISSING",
                "Current Price": 100,
                "Target 1": None,
                "Relative Strength": None,
                "Score": 30,
            },
        ]
    )

    page = _generate_html(
        dataframe,
        "12 August 2026, 10:00 PM",
        page_key="technical",
    )
    top_table = page.split('<table id="topTable">', 1)[1].split("</table>", 1)[0]
    ordered_symbols = [
        "STRONGER_REC",
        "STRONGER_SIGNAL",
        "STRONGER_TREND",
        "HIGHER_RS",
        "LARGER_TIE_GAP",
        "BASE",
        "MISSING",
    ]

    positions = [top_table.index(f">{symbol}</span>") for symbol in ordered_symbols]
    assert positions == sorted(positions)
    assert ">LARGEST_GAP</span>" not in page
    assert ">LARGEST_GAP</span>" not in page


def test_technical_price_levels_show_direction_from_current_price():
    dataframe = report.prepare_results_dataframe(
        [
            {
                "Symbol": "LEVELS",
                "Current Price": 100,
                "Target 1": 110,
                "Target 2": 120,
                "Target 3": 130,
                "20 MA": 90,
                "50 MA": 95,
                "200 MA": 100,
                "Stop Loss": 80,
                "Score": 90,
            },
        ]
    )

    page = _generate_html(
        dataframe,
        "13 August 2026, 12:00 PM EDT",
        page_key="technical",
    )

    assert page.count('class="target-arrow-up" title="Above current price"') == 6
    assert page.count('class="target-arrow-down" title="Below current price"') == 6
    for value in (80, 90, 95, 100, 110, 120, 130):
        assert page.count(f'data-price-level="{float(value)}"') == 2
    assert "row.querySelectorAll('[data-price-level]')" in page


def test_analyst_price_levels_show_direction_from_current_price():
    dataframe = report.prepare_results_dataframe(
        [
            {
                "Symbol": "LEVELS",
                "Current Price": 100,
                "Support Low": 90,
                "Support High": 110,
                "Resistance Low": 95,
                "Resistance High": 105,
                "Score": 90,
            }
        ]
    )

    page = _generate_html(
        dataframe,
        "13 August 2026, 02:00 PM EDT",
        page_key="analysts",
    )

    for value in ("$90.00", "$95.00"):
        assert page.count(
            f'{value} <span class="target-arrow-down" '
            'title="Below current price">&#8595;</span>'
        ) == 2
    for value in ("$105.00", "$110.00"):
        assert page.count(
            f'{value} <span class="target-arrow-up" '
            'title="Above current price">&#8593;</span>'
        ) == 2


def test_kpi_chart_normalizes_and_equal_weights_at_common_start():
    dates = pd.date_range("2026-01-02", periods=3, freq="D")
    histories = {
        "^GSPC": pd.DataFrame({"Close": [100, 110, 120]}, index=dates),
        "AAA": pd.DataFrame({"Close": [10, 20, 30]}, index=dates),
        "BBB": pd.DataFrame({"Close": [20, 20, 40]}, index=dates),
    }

    chart_data = _build_kpi_chart_data(
        pd.DataFrame({"Symbol": ["AAA", "BBB", "AAA"]}),
        history_loader=lambda symbol, period: histories[symbol],
        now=datetime(2026, 1, 10, tzinfo=ZoneInfo("America/New_York")),
    )

    assert chart_data["labels"] == ["2026-01-02", "2026-01-03", "2026-01-04"]
    assert chart_data["sp500"] == [100.0, 110.0, 120.0]
    assert chart_data["top20"] == [100.0, 150.0, 250.0]
    assert chart_data["constituents"] == ["AAA", "BBB"]
    assert chart_data["selected_count"] == 2


def test_kpi_chart_aligns_dates_and_skips_invalid_constituents(capsys):
    histories = {
        "^GSPC": pd.DataFrame(
            {"Close": [100, 110, 120]},
            index=pd.to_datetime(["2026-01-02", "2026-01-03", "2026-01-04"]),
        ),
        "AAA": pd.DataFrame(
            {"Close": [10, 20, 30]},
            index=pd.to_datetime(["2026-01-02", "2026-01-03", "2026-01-04"]),
        ),
        "LATE": pd.DataFrame(
            {"Close": [40, 80]},
            index=pd.to_datetime(["2026-01-03", "2026-01-04"]),
        ),
        "BAD": pd.DataFrame(
            {"Close": ["not-a-price"]},
            index=pd.to_datetime(["2026-01-03"]),
        ),
    }

    chart_data = _build_kpi_chart_data(
        pd.DataFrame({"Symbol": ["AAA", "BAD", "LATE"]}),
        history_loader=lambda symbol, period: histories[symbol],
        now=datetime(2026, 1, 10, tzinfo=ZoneInfo("America/New_York")),
    )

    assert chart_data["labels"] == ["2026-01-03", "2026-01-04"]
    assert chart_data["sp500"] == [100.0, 109.0909]
    assert chart_data["top20"] == [100.0, 175.0]
    assert chart_data["constituents"] == ["AAA", "LATE"]
    assert "KPI performance history unavailable for BAD" in capsys.readouterr().err


def test_kpi_chart_renders_exactly_two_indexed_line_series():
    dataframe = pd.DataFrame(
        [
            {
                "Symbol": "AAA",
                "Score": 90,
                "Entry": 100,
                "Current Price": 110,
                "Support Low": 95,
                "Resistance Low": 120,
            }
        ]
    )
    page = _generate_html(
        dataframe,
        "15 August 2026, 10:00 AM EDT",
        page_key="landing",
        chart_data={
            "labels": ["2026-08-13", "2026-08-14"],
            "sp500": [100.0, 101.0],
            "top20": [100.0, 102.0],
            "constituents": ["AAA"],
            "selected_count": 1,
        },
    )

    assert '<canvas id="performanceChart"' in page
    assert "type: 'line'" in page
    assert page.count("label: 'S&P 500'") == 1
    assert page.count("label: 'Equal-weight Top 20'") == 1
    assert "Indexed to 100" in page
    assert "not raw dollars" in page
    assert "Recommendation Breakdown" not in page
    assert "recChart" not in page
    assert "Top 20 Qualifying Stocks: Suggested Entry vs Current Price" in page
    assert 'id="top20DetailsTable"' in page
    assert (
        'data-symbol="AAA" data-entry="100.0" data-support-low="95.0" '
        'data-resistance-low="120.0"'
        in page
    )
    assert '<span class="detail-current">($110.00)</span>' in page
    assert (
        "<th>Symbol</th><th>Suggested Entry</th>"
        "<th>Support Low</th><th>Resistance Low</th>"
        "<th>Difference</th><th>Change %</th>"
    ) in page
    assert "<th>Support Low</th><th>Resistance Low</th>" in page
    assert '<td class="detail-support-low">$95.00</td>' in page
    assert '<td class="detail-resistance-low">$120.00</td>' in page
    assert '<td class="detail-difference price-gain">+10.00</td>' in page
    assert '<td class="detail-percent price-gain">+10.00%</td>' in page
    assert "Entry is the scanner suggestion, not an actual purchase price." in page
    assert "Symbols at or above Resistance Low are excluded." in page
    assert "current >= resistanceLow" in page
    assert "row.remove();" in page
    assert "current < supportLow" in page
    assert "current < entry" in page
    assert "updateTop20Details(snapshot.prices)" in page


def test_kpi_chart_renders_unavailable_message_without_javascript_data():
    page = _generate_html(
        pd.DataFrame([{"Symbol": "AAA", "Score": 90}]),
        "15 August 2026, 10:00 AM EDT",
        page_key="landing",
        chart_data=None,
    )

    assert "Performance chart unavailable" in page
    assert '<canvas id="performanceChart"' not in page
    assert "const chartData = null;" in page


def test_top_twenty_details_handles_missing_entry_without_false_change():
    page = _generate_html(
        pd.DataFrame(
            [{"Symbol": "MISSING", "Score": 90, "Current Price": 50}]
        ),
        "15 August 2026, 10:00 AM EDT",
        page_key="landing",
        chart_data=None,
    )

    assert (
        'data-symbol="MISSING" data-entry="" data-support-low="" '
        'data-resistance-low=""'
        in page
    )
    assert '<td class="detail-entry">Unavailable</td>' in page
    assert '<td class="detail-support-low">Unavailable</td>' in page
    assert '<td class="detail-resistance-low">Unavailable</td>' in page
    assert '<td class="detail-difference ">Unavailable</td>' in page
    assert '<td class="detail-percent ">Unavailable</td>' in page


def test_top_twenty_details_excludes_prices_at_or_above_resistance_low():
    details = html_report._build_top_twenty_details(
        pd.DataFrame(
            [
                {
                    "Symbol": "BELOW",
                    "Current Price": 99,
                    "Resistance Low": 100,
                },
                {
                    "Symbol": "EQUAL",
                    "Current Price": 100,
                    "Resistance Low": 100,
                },
                {
                    "Symbol": "ABOVE",
                    "Current Price": 101,
                    "Resistance Low": 100,
                },
                {
                    "Symbol": "MISSING",
                    "Current Price": 101,
                    "Resistance Low": None,
                },
            ]
        )
    )

    assert [detail["symbol"] for detail in details] == ["BELOW", "MISSING"]


def test_top_twenty_details_backfills_from_full_ranked_scan():
    rows = [
        {
            "Symbol": f"EXCLUDED{index}",
            "Current Price": 100,
            "Resistance Low": 100,
        }
        for index in range(15)
    ]
    rows.extend(
        {
            "Symbol": f"QUALIFIED{index}",
            "Current Price": 99,
            "Resistance Low": 100,
        }
        for index in range(25)
    )

    details = html_report._build_top_twenty_details(pd.DataFrame(rows))

    assert len(details) == 20
    assert details[0]["symbol"] == "QUALIFIED0"
    assert details[-1]["symbol"] == "QUALIFIED19"


def test_top_twenty_details_highlights_entry_and_below_support_opportunities():
    page = _generate_html(
        pd.DataFrame(
            [
                {
                    "Symbol": "OPPORTUNITY",
                    "Entry": 90,
                    "Current Price": 80,
                    "Support Low": 85,
                    "Resistance Low": 100,
                }
            ]
        ),
        "15 August 2026, 10:00 AM EDT",
        page_key="landing",
        chart_data=None,
    )

    assert '<td class="detail-entry entry-opportunity">$90.00</td>' in page
    assert '<span class="detail-current below-support">($80.00)</span>' in page
    assert ".below-support { color: #66bb6a;" in page
    assert ".entry-opportunity { background: #1b5e20;" in page


def test_landing_kpis_open_filterable_recommendation_stock_lists():
    page = _generate_html(
        pd.DataFrame(
            [
                {"Rank": 1, "Symbol": "AAA", "Recommendation": "Strong Buy", "Score": 90},
                {"Rank": 2, "Symbol": "BBB", "Recommendation": "Buy", "Score": 80},
                {"Rank": 3, "Symbol": "CCC", "Recommendation": "Watch", "Score": 60},
            ]
        ),
        "28 August 2026, 8:00 AM EDT",
        page_key="landing",
        chart_data=None,
    )

    for category in ("all", "strong-buy", "buy", "accumulate", "watch", "avoid"):
        assert f'data-kpi-filter="{category}"' in page
    assert 'id="kpiDetails"' in page
    assert 'id="kpiStockSearch"' in page
    assert '<tr data-kpi-category="strong-buy">' in page
    assert '<tr data-kpi-category="buy">' in page
    assert '<tr data-kpi-category="watch">' in page
    assert "initializeKpiDrilldown();" in page
    assert "function filterKpiStockRows()" in page


def test_html_export_creates_three_stable_linked_pages(tmp_path, monkeypatch):
    monkeypatch.setattr(html_report, "REPORT_FOLDER", str(tmp_path))
    snapshot_path = tmp_path / "prices.json"
    monkeypatch.setattr(html_report, "PRICE_SNAPSHOT_PATH", snapshot_path)
    monkeypatch.setattr(html_report, "_build_kpi_chart_data", lambda *args, **kwargs: None)

    archived_page = html_report.export_html_report(
        [{"Symbol": "ABC", "Score": 90, "Current Price": 10}],
        quiet=True,
    )

    date_folder = Path(archived_page).parent
    assert (date_folder / "landing.html").exists()
    assert (date_folder / "technical.html").exists()
    assert (date_folder / "analysts.html").exists()
    assert (date_folder / "bought-selection.html").exists()
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    assert snapshot["source"] == "full_scan"
    assert snapshot["prices"] == {"ABC": 10.0}


def test_html_export_survives_results_without_snapshot_prices(
    tmp_path, monkeypatch, capsys
):
    monkeypatch.setattr(html_report, "REPORT_FOLDER", str(tmp_path))
    monkeypatch.setattr(
        html_report,
        "PRICE_SNAPSHOT_PATH",
        tmp_path / "prices.json",
    )
    monkeypatch.setattr(html_report, "_build_kpi_chart_data", lambda *args, **kwargs: None)

    archived_page = html_report.export_html_report(
        [{"Symbol": "ABC", "Score": 90, "Current Price": None}],
        quiet=True,
    )

    assert Path(archived_page).exists()
    assert not (tmp_path / "prices.json").exists()
    assert "Price snapshot was not updated" in capsys.readouterr().err


def test_setup_priority_orders_supported_entry_setups():
    assert setup_priority("Strong Uptrend") > setup_priority("Pullback to 20 MA")
    assert setup_priority("Pullback to 20 MA") > setup_priority("Pullback to 50 MA")
    assert setup_priority("Pullback to 50 MA") > setup_priority("Breakout Candidate")
    assert setup_priority("Breakout Candidate") > setup_priority("Neutral")


def test_setup_priority_is_secondary_to_score(monkeypatch):
    monkeypatch.setattr("stockscanner.ranking.load_exceptions", lambda: set())
    results = [
        {
            "Symbol": "LOW",
            "Score": 89,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 10,
            "Relative Strength": 50,
        },
        {
            "Symbol": "BREAK",
            "Score": 90,
            "Signal": "Breakout Candidate",
            "Risk/Reward": 1,
            "Relative Strength": 5,
        },
        {
            "Symbol": "PULL50",
            "Score": 90,
            "Signal": "Pullback to 50 MA",
            "Risk/Reward": 1,
            "Relative Strength": 5,
        },
        {
            "Symbol": "PULL20",
            "Score": 90,
            "Signal": "Pullback to 20 MA",
            "Risk/Reward": 1,
            "Relative Strength": 5,
        },
        {
            "Symbol": "UP",
            "Score": 90,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 1,
            "Relative Strength": 5,
        },
    ]

    ranked = rank_stocks(results)
    prepared = report.prepare_results_dataframe(results)
    expected = ["UP", "PULL20", "PULL50", "BREAK", "LOW"]

    assert ranked["Symbol"].tolist() == expected
    assert prepared["Symbol"].tolist() == expected


def test_analyst_data_calculates_target_upside_and_uses_cache(tmp_path, monkeypatch):
    calls = []

    class FakeTicker:
        def get_info(self):
            calls.append("requested")
            return {
                "recommendationKey": "buy",
                "targetMeanPrice": 125,
            }

    monkeypatch.setattr(
        "stockscanner.analyst_data.yf.Ticker", lambda symbol: FakeTicker()
    )

    first = get_analyst_data("ABC", 100, cache_directory=str(tmp_path))
    second = get_analyst_data("ABC", 100, cache_directory=str(tmp_path))

    assert first == {"Analyst Rating": "Buy", "Target Upside": 25.0}
    assert second == first
    assert calls == ["requested"]


def test_analyst_data_only_breaks_complete_technical_ties(monkeypatch):
    monkeypatch.setattr("stockscanner.ranking.load_exceptions", lambda: set())
    results = [
        {
            "Symbol": "BETTER_TECHNICAL",
            "Score": 90,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 2,
            "Relative Strength": 20,
            "Analyst Rating": "Hold",
            "Target Upside": 5,
        },
        {
            "Symbol": "STRONG_BUY",
            "Score": 90,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 1,
            "Relative Strength": 20,
            "Analyst Rating": "Strong Buy",
            "Target Upside": 30,
        },
        {
            "Symbol": "BUY_HIGH_TARGET",
            "Score": 90,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 1,
            "Relative Strength": 20,
            "Analyst Rating": "Buy",
            "Target Upside": 25,
        },
        {
            "Symbol": "BUY_LOW_TARGET",
            "Score": 90,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 1,
            "Relative Strength": 20,
            "Analyst Rating": "Buy",
            "Target Upside": 10,
        },
    ]

    ranked = rank_stocks(results)

    assert ranked["Symbol"].tolist() == [
        "BETTER_TECHNICAL",
        "STRONG_BUY",
        "BUY_HIGH_TARGET",
        "BUY_LOW_TARGET",
    ]
    assert analyst_rating_priority("Strong Buy") > analyst_rating_priority("Buy")


def test_scan_result_places_analyst_and_risk_columns_after_sector(monkeypatch):
    history = pd.DataFrame(
        {
            "Close": [10.0] * 200,
            "High": [11.0] * 200,
            "Volume": [1_000_000] * 200,
            "MA20": [9.5] * 200,
            "MA50": [9.0] * 200,
            "MA200": [8.0] * 200,
            "RSI": [60.0] * 200,
            "MACD": [1.0] * 200,
            "MACD_SIGNAL": [0.5] * 200,
            "AVG_VOLUME": [900_000] * 200,
        }
    )
    monkeypatch.setattr("stockscanner.scan.download_data", lambda symbol: history)
    monkeypatch.setattr("stockscanner.scan.calculate_indicators", lambda data: data)
    monkeypatch.setattr("stockscanner.scan.calculate_relative_strength", lambda symbol: 10)
    monkeypatch.setattr("stockscanner.scan.download_intraday_snapshot", lambda symbol: None)
    monkeypatch.setattr("stockscanner.scan.score_stock", lambda data, strength: 90)
    monkeypatch.setattr("stockscanner.scan.generate_signal", lambda data: "Strong Uptrend")
    monkeypatch.setattr(
        "stockscanner.scan.generate_trade_plan",
        lambda data, available_cash, risk_percent: {
            "Trend": "Strong Uptrend",
            "Entry": 10,
            "Stop": 9,
            "Target1": 11,
            "Target2": 12,
            "Target3": 13,
            "RR": 2,
            "Shares": 25,
            "Investment": 250,
        },
    )
    monkeypatch.setattr(
        "stockscanner.scan.get_analyst_data",
        lambda symbol, current_price: {
            "Analyst Rating": "Buy",
            "Target Upside": 20,
        },
    )

    result = process_stock(
        {"Symbol": "ABC", "Market": "NYSE", "Sector": "Technology"},
        quiet=True,
    )
    columns = list(result)

    sector_index = columns.index("Sector")
    assert columns[sector_index : sector_index + 6] == [
        "Sector",
        "Analyst Rating",
        "Target Upside",
        "Suggested Shares",
        "Risk/Reward",
        "Priority",
    ]


def test_completed_daily_data_drops_current_session_candle():
    index = pd.to_datetime(["2026-08-11", "2026-08-12"]).tz_localize(
        "America/New_York"
    )
    history = pd.DataFrame({"Close": [10, 11]}, index=index)
    now = datetime(2026, 8, 12, 12, tzinfo=ZoneInfo("America/New_York"))

    completed = completed_daily_data(history, now=now)

    assert completed["Close"].tolist() == [10]


def test_intraday_snapshot_uses_latest_current_session_quote(monkeypatch):
    index = pd.to_datetime(
        ["2026-08-12 09:00", "2026-08-12 10:15"]
    ).tz_localize("America/New_York")
    intraday = pd.DataFrame(
        {"Close": [10.0, 10.5], "Volume": [100, 250]},
        index=index,
    )

    class FakeTicker:
        def history(self, **kwargs):
            assert kwargs == {"period": "1d", "interval": "1m", "prepost": True}
            return intraday

    monkeypatch.setattr(
        "stockscanner.market_data.yf.Ticker", lambda symbol: FakeTicker()
    )
    now = datetime(2026, 8, 12, 10, 16, tzinfo=ZoneInfo("America/New_York"))

    snapshot = download_intraday_snapshot("ABC", now=now)

    assert snapshot["price"] == 10.5
    assert snapshot["volume"] == 350
    assert snapshot["timestamp"].startswith("2026-08-12T10:15:00")


def test_process_stock_overlays_intraday_price_after_daily_indicators(monkeypatch):
    index = pd.date_range(
        "2025-10-01",
        periods=200,
        freq="B",
        tz="America/New_York",
    )
    history = pd.DataFrame(
        {
            "Close": [10.0] * 200,
            "High": [11.0] * 200,
            "Volume": [1_000_000] * 200,
            "MA20": [9.5] * 200,
            "MA50": [9.0] * 200,
            "MA200": [8.0] * 200,
            "RSI": [60.0] * 200,
            "MACD": [1.0] * 200,
            "MACD_SIGNAL": [0.5] * 200,
            "AVG_VOLUME": [900_000] * 200,
        },
        index=index,
    )
    observed = {}
    monkeypatch.setattr("stockscanner.scan.download_data", lambda symbol: history)
    monkeypatch.setattr("stockscanner.scan.completed_daily_data", lambda data: data)
    monkeypatch.setattr("stockscanner.scan.calculate_indicators", lambda data: data)
    monkeypatch.setattr(
        "stockscanner.scan.download_intraday_snapshot",
        lambda symbol: {
            "price": 12.5,
            "volume": 1_500_000,
            "timestamp": "2026-08-12T10:15:00-04:00",
        },
    )
    monkeypatch.setattr("stockscanner.scan.calculate_relative_strength", lambda symbol: 10)
    monkeypatch.setattr(
        "stockscanner.scan.analyze_support_resistance",
        lambda data, current_price: observed.setdefault(
            "zone_input", (data["Close"].iloc[-1], current_price)
        )
        and {
            "Zone Status": "Between Zones",
            "Support Low": 9,
            "Support High": 10,
            "Resistance Low": 13,
            "Resistance High": 14,
        },
    )
    monkeypatch.setattr(
        "stockscanner.scan.score_stock",
        lambda data, strength: observed.setdefault(
            "score_input", (data["Close"].iloc[-1], data["Volume"].iloc[-1])
        )
        and 90,
    )
    monkeypatch.setattr("stockscanner.scan.generate_signal", lambda data: "Strong Uptrend")
    monkeypatch.setattr(
        "stockscanner.scan.generate_trade_plan",
        lambda data, available_cash, risk_percent: {
            "Trend": "Strong Uptrend",
            "Entry": 10,
            "Stop": 9,
            "Target1": 11,
            "Target2": 12,
            "Target3": 13,
            "RR": 2,
            "Shares": 25,
            "Investment": 250,
        },
    )
    monkeypatch.setattr(
        "stockscanner.scan.get_analyst_data",
        lambda symbol, current_price: {
            "Analyst Rating": "Buy",
            "Target Upside": 20,
        },
    )

    result = process_stock(
        {"Symbol": "ABC", "Market": "NYSE", "Sector": "Technology"},
        quiet=True,
    )

    assert observed["score_input"] == (12.5, 1_500_000)
    assert observed["zone_input"] == (10.0, 12.5)
    assert result["Current Price"] == 12.5
    assert result["Zone Status"] == "Between Zones"
    assert result["Price As Of"] == "2026-08-12T10:15:00-04:00"
