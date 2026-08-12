import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from stockscanner import report
from stockscanner.html_report import _generate_html
from stockscanner.ranking import rank_stocks
from stockscanner.support_resistance import analyze_support_resistance


def _oscillating_history(last_close=None):
    periods = 100
    x = np.arange(periods)
    close = 100 + 5 * np.sin(2 * np.pi * x / 10)
    dataframe = pd.DataFrame(
        {
            "Close": close,
            "High": close + 1,
            "Low": close - 1,
            "Volume": np.full(periods, 1_000.0),
        }
    )
    if last_close is not None:
        dataframe.loc[periods - 1, ["Close", "High", "Low"]] = [
            last_close,
            last_close + 1,
            last_close - 1,
        ]
    dataframe["MA20"] = dataframe["Close"].rolling(20).mean()
    dataframe["MA50"] = dataframe["Close"].rolling(50).mean()
    dataframe["MA200"] = np.nan
    return dataframe


@pytest.mark.parametrize(
    ("price", "expected_status"),
    [
        (94, "At Support"),
        (100, "At Resistance"),
        (96, "Between Zones"),
        (104, "Breakout Above Resistance"),
        (88, "Breakdown Below Support"),
    ],
)
def test_zone_statuses(price, expected_status):
    analysis = analyze_support_resistance(_oscillating_history(), price)

    assert analysis["Zone Status"] == expected_status


def test_broken_resistance_becomes_support():
    analysis = analyze_support_resistance(
        _oscillating_history(last_close=110),
        current_price=106,
    )

    assert analysis["Zone Status"] == "At Support"
    assert "role reversal" in analysis["Support Details"]
    assert "swing high" in analysis["Support Details"]


def test_broken_support_becomes_resistance():
    analysis = analyze_support_resistance(
        _oscillating_history(last_close=90),
        current_price=94,
    )

    assert analysis["Zone Status"] == "At Resistance"
    assert "role reversal" in analysis["Resistance Details"]
    assert "swing low" in analysis["Resistance Details"]


def test_atr_sets_zone_tolerance_and_zero_atr_uses_percentage_fallback():
    atr_data = pd.DataFrame(
        {
            "Close": [100.0] * 30,
            "High": [102.0] * 30,
            "Low": [98.0] * 30,
            "Volume": [1_000.0] * 30,
        }
    )
    flat_data = atr_data.assign(High=100.0, Low=100.0)

    atr_analysis = analyze_support_resistance(atr_data)
    fallback_analysis = analyze_support_resistance(flat_data)

    assert atr_analysis["Zone Tolerance"] == 2.0
    assert atr_analysis["Zone Tolerance %"] == 2.0
    assert fallback_analysis["Zone Tolerance"] == 1.0


def test_insufficient_or_missing_data_is_explicitly_unavailable():
    short_data = _oscillating_history().head(10)
    missing_low = _oscillating_history().drop(columns=["Low"])

    assert analyze_support_resistance(short_data)["Zone Status"] == "Unavailable"
    assert analyze_support_resistance(missing_low)["Zone Status"] == "Unavailable"
    assert analyze_support_resistance(None)["Support Confidence"] == "Unavailable"


def test_zone_fields_do_not_change_ranking_hierarchy(monkeypatch):
    monkeypatch.setattr("stockscanner.ranking.load_exceptions", lambda: set())
    base_results = [
        {
            "Symbol": "FIRST",
            "Score": 90,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 2,
            "Relative Strength": 20,
        },
        {
            "Symbol": "SECOND",
            "Score": 90,
            "Signal": "Pullback to 20 MA",
            "Risk/Reward": 3,
            "Relative Strength": 30,
        },
        {
            "Symbol": "THIRD",
            "Score": 80,
            "Signal": "Strong Uptrend",
            "Risk/Reward": 10,
            "Relative Strength": 50,
        },
    ]
    with_zones = [
        {
            **row,
            "Zone Status": status,
            "Support Confidence": confidence,
        }
        for row, status, confidence in zip(
            base_results,
            ["Breakdown Below Support", "At Support", "Breakout Above Resistance"],
            ["Low", "High", "High"],
        )
    ]

    expected = rank_stocks(base_results)["Symbol"].tolist()

    assert rank_stocks(with_zones)["Symbol"].tolist() == expected
    assert report.prepare_results_dataframe(with_zones)["Symbol"].tolist() == expected


def test_report_places_and_formats_zone_fields(tmp_path):
    dataframe = report.prepare_results_dataframe(
        [
            {
                "Symbol": "ABC",
                "Current Price": 100,
                "Score": 90,
                "Zone Status": "At Support",
                "Support Low": "98.5",
                "Support High": "100.5",
                "Support Distance %": "0.5",
            }
        ]
    )

    current_price_index = dataframe.columns.get_loc("Current Price")
    assert dataframe.columns[current_price_index + 1] == "Zone Status"
    assert dataframe["Support Low"].iloc[0] == 98.5

    html = _generate_html(dataframe, "12 August 2026, 07:00 PM")
    assert "$98.50" in html
    assert "0.50%" in html

    workbook_path = tmp_path / "zones.xlsx"
    report.export_excel_workbook(workbook_path, dataframe)
    worksheet = load_workbook(workbook_path)["Complete Scan"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    assert worksheet.cell(2, headers["Support Low"]).number_format == "$#,##0.00"
    assert worksheet.cell(2, headers["Support Distance %"]).number_format == '0.00"%"'
